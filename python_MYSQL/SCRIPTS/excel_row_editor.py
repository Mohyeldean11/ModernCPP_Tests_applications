from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


def history_path_for(excel_path: Path) -> Path:
    return excel_path.with_name(f".{excel_path.stem}_edit_history.jsonl")


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def load_history(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    rows: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                payload = json.loads(line)
                if isinstance(payload, dict):
                    rows.append(payload)
            except json.JSONDecodeError:
                continue
    return rows


def append_history(path: Path, payload: Dict[str, Any]) -> None:
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(payload, ensure_ascii=False) + "\n")


def choose_sheet(workbook, preferred: Optional[str]) -> str:
    names = workbook.sheetnames
    if preferred and preferred in names:
        return preferred

    print("\nAvailable sheets:")
    for idx, name in enumerate(names, start=1):
        print(f"  {idx}. {name}")

    while True:
        raw = input("Choose sheet number (Enter for first sheet): ").strip()
        if not raw:
            return names[0]
        if raw.isdigit():
            i = int(raw)
            if 1 <= i <= len(names):
                return names[i - 1]
        print("Invalid selection.")


def build_row_dict(headers: List[str], row_values: List[Any]) -> Dict[str, Any]:
    data: Dict[str, Any] = {}
    for idx, value in enumerate(row_values):
        key = headers[idx] if idx < len(headers) and headers[idx] else f"Column_{idx + 1}"
        data[str(key)] = value
    return data


def find_matching_rows(worksheet, search_value: str) -> List[int]:
    needle = search_value.casefold()
    matches: List[int] = []
    for row_idx in range(2, worksheet.max_row + 1):
        row_has_match = False
        for col_idx in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            if needle in str(value).casefold():
                row_has_match = True
                break
        if row_has_match:
            matches.append(row_idx)
    return matches


def parse_with_type(original: Any, new_text: str) -> Any:
    if new_text == "":
        return None
    if original is None:
        return new_text
    if isinstance(original, bool):
        lowered = new_text.strip().lower()
        if lowered in {"true", "1", "yes", "y"}:
            return True
        if lowered in {"false", "0", "no", "n"}:
            return False
        return new_text
    if isinstance(original, int) and not isinstance(original, bool):
        try:
            return int(new_text)
        except ValueError:
            return new_text
    if isinstance(original, float):
        try:
            return float(new_text)
        except ValueError:
            return new_text
    return new_text


def pick_row(matches: List[int], worksheet, headers: List[str]) -> Optional[int]:
    print("\nMatching rows:")
    for idx, row_num in enumerate(matches, start=1):
        values = [worksheet.cell(row=row_num, column=col).value for col in range(1, worksheet.max_column + 1)]
        row_map = build_row_dict(headers, values)
        print(f"  {idx}. Excel row {row_num}: {row_map}")

    while True:
        raw = input("Select row number to edit (or Enter to cancel): ").strip()
        if raw == "":
            return None
        if raw.isdigit():
            i = int(raw)
            if 1 <= i <= len(matches):
                return matches[i - 1]
        print("Invalid row selection.")


def pick_column(headers: List[str], row_values: List[Any]) -> Optional[int]:
    print("\nColumns in selected row:")
    for idx, value in enumerate(row_values, start=1):
        header = headers[idx - 1] if idx - 1 < len(headers) else ""
        title = header if header else f"Column_{idx}"
        print(f"  {idx}. {title} = {value}")

    while True:
        raw = input("Choose column number to modify (or Enter to cancel): ").strip()
        if raw == "":
            return None
        if raw.isdigit():
            i = int(raw)
            if 1 <= i <= len(row_values):
                return i
        print("Invalid column selection.")


def show_history(history_file: Path, excel_path: Path, sheet_name: str) -> None:
    rows = load_history(history_file)
    filtered = [
        row
        for row in rows
        if row.get("excel_path") == str(excel_path.resolve()) and row.get("sheet") == sheet_name
    ]

    if not filtered:
        print("\nNo history entries found for this sheet.")
        return

    print(f"\nHistory for '{sheet_name}' ({len(filtered)} entries):")
    for idx, entry in enumerate(filtered, start=1):
        print(
            f"{idx}. {entry.get('timestamp')} | row {entry.get('row')} | "
            f"{entry.get('column')} ({entry.get('header')}) | "
            f"{entry.get('old_value')} -> {entry.get('new_value')} | "
            f"search='{entry.get('search_value')}'"
        )


def run_editor(excel_path: Path, initial_sheet: Optional[str]) -> None:
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    workbook = load_workbook(excel_path)
    sheet_name = choose_sheet(workbook, initial_sheet)
    history_file = history_path_for(excel_path)

    while True:
        worksheet = workbook[sheet_name]
        headers = [
            worksheet.cell(row=1, column=col).value if worksheet.max_row >= 1 else None
            for col in range(1, worksheet.max_column + 1)
        ]
        headers = ["" if h is None else str(h) for h in headers]

        print("\nMenu:")
        print("  1. Search value and edit matching row")
        print("  2. View change history for current sheet")
        print("  3. Switch sheet")
        print("  4. Exit")

        choice = input("Choose an option: ").strip()

        if choice == "1":
            search_value = input("Enter value to search for: ").strip()
            if not search_value:
                print("Search value is required.")
                continue

            matches = find_matching_rows(worksheet, search_value)
            if not matches:
                print("No matching rows found.")
                continue

            row_num = pick_row(matches, worksheet, headers)
            if row_num is None:
                continue

            row_values = [worksheet.cell(row=row_num, column=col).value for col in range(1, worksheet.max_column + 1)]
            col_num = pick_column(headers, row_values)
            if col_num is None:
                continue

            old_value = worksheet.cell(row=row_num, column=col_num).value
            header = headers[col_num - 1] if col_num - 1 < len(headers) and headers[col_num - 1] else f"Column_{col_num}"
            new_text = input(f"Enter new value for '{header}' (empty = clear cell): ")
            new_value = parse_with_type(old_value, new_text)

            worksheet.cell(row=row_num, column=col_num).value = new_value
            workbook.save(excel_path)

            entry = {
                "timestamp": now_iso(),
                "excel_path": str(excel_path.resolve()),
                "sheet": sheet_name,
                "search_value": search_value,
                "row": row_num,
                "column": col_num,
                "header": header,
                "old_value": old_value,
                "new_value": new_value,
            }
            append_history(history_file, entry)

            print("Update saved to the same Excel file.")

        elif choice == "2":
            show_history(history_file, excel_path, sheet_name)

        elif choice == "3":
            sheet_name = choose_sheet(workbook, None)

        elif choice == "4":
            print("Exiting.")
            break

        else:
            print("Invalid menu option.")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Search rows in an Excel sheet, edit a selected value, save changes in place, "
            "and keep a local change history."
        )
    )
    parser.add_argument("excel_file", nargs="?", help="Path to the .xlsx file")
    parser.add_argument("--sheet", help="Sheet name to open first")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    excel_input = args.excel_file or input("Enter path to Excel file (.xlsx): ").strip()
    if not excel_input:
        raise SystemExit("Excel file path is required.")

    excel_path = Path(excel_input).expanduser().resolve()
    run_editor(excel_path, args.sheet)


if __name__ == "__main__":
    main()
