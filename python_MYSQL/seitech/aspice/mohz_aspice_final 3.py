#!/usr/bin/env python3
"""
Unified ASPICE Requirements Analyzer - Final Version
Supports both SYS.2 (System Requirements) and SWE.1 (Software Requirements) analysis.

Key Features:
1. Comprehensive structural analysis for both requirement types
2. Intelligent attribute-BP mapping 
3. LLM-focused compliance checking (70% AI, 30% rule-based)
4. Multi-sheet Excel output with detailed analysis
5. Smart recommendation engine
6. Quality checklist assessment (5-point)
7. Choice between SYS.2 and SWE.1 analysis modes
8. Interactive menu system
9. Analysis-type-specific configuration files
10. Robust encoding detection for CSV files
11. Input files always requested from user (not from config)
12. SWE.1 now uses BP1-BP4 only, includes ASIL attribute
"""

import argparse
import json
import requests
import re
import os
from typing import Dict, List, Tuple, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import Counter
from dataclasses import dataclass
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class StructuralAnalysis:
    """File-level structural analysis results"""
    has_groups: bool
    group_count: int
    groups_detected: List[str]
    functional_requirements: int
    non_functional_requirements: int
    undefined_type_requirements: int
    total_requirements: int
    functional_coverage: float
    structural_completeness_score: float
    atomic_requirements_ratio: float = 0.0

@dataclass
class RequirementAnalysis:
    """Individual requirement analysis results"""
    requirement_id: str
    is_compliant: bool
    overall_confidence: float
    bp_scores: Dict[str, float]
    violated_practices: List[str]
    recommendations: List[str]
    # is_consistent: bool
    # is_testable: bool
    missing_attributes: List[str]
    structural_impact: Dict[str, str]
    quality_checklist: Dict[str, bool] = None
    is_atomic: bool = True
    
    def __post_init__(self):
        """Ensure quality_checklist always exists with default values"""
        if self.quality_checklist is None:
            self.quality_checklist = {
                'Complete': False,
                'Correct': False,
                'Feasible': False,
                'Verifiable': False,
                'Atomic': False
            }

# SYS.2 Base Practices (BP1-BP4 only)
SYS2_BASE_PRACTICES = {
    "BP1": {
        "name": "Specify system requirements",
        "description": "Use stakeholder requirements to identify and document functional and non-functional system requirements in a clear, verifiable, and complete manner.",
        "focus_area": "Requirements specification quality and completeness",
        "key_attributes": ["ID", "Text", "Object_Type", "ASIL"],
        "llm_focus": "Assess requirement clarity, completeness, and specification quality from available text and attributes."
    },
    "BP2": {
        "name": "Structure and prioritize system requirements",
        "description": "Organize requirements into logical groups, ensure unique identification, and prioritize based on stakeholder needs.",
        "focus_area": "Requirements organization and prioritization",
        "key_attributes": ["ASIL", "Object_Type", "Group_Structure"],
        "llm_focus": "Evaluate organizational structure, logical grouping, and systematic approach from ID patterns and available data."
    },
    "BP3": {
        "name": "Analyze system requirements",
        "description": "Evaluate requirements for correctness, completeness, and technical feasibility with proper verification approach.",
        "focus_area": "Requirements analysis and verification planning",
        "key_attributes": ["Verification_Strategy", "Verification_Criteria", "Review_Status"],
        "llm_focus": "Assess verification planning, analysis depth, and requirement quality from verification data and content."
    },
    "BP4": {
        "name": "Analyze impact on system context",
        "description": "Assess how requirements affect interfaces, operating environment, and system elements.",
        "focus_area": "System context and environmental impact analysis",
        "key_attributes": ["Text", "Verification_Criteria", "ASIL", "System_Feature"],
        "llm_focus": "Identify system context considerations, interface impacts, environmental factors, and system feature mapping from requirement text."
    }
}

# SWE.1 Base Practices (BP1-BP4 only - removed BP5 and BP6)
SWE1_BASE_PRACTICES = {
    "BP1": {
        "name": "Specify software requirements",
        "description": "Use system requirements and architecture to identify and document functional and non-functional software requirements clearly and completely.",
        "focus_area": "Software requirements specification quality and completeness",
        "key_attributes": ["ID", "Text", "Object_Type"],
        "llm_focus": "Assess software requirement clarity, completeness, specification quality, and alignment with system requirements from available text and attributes."
    },
    "BP2": {
        "name": "Structure software requirements", 
        "description": "Structure and prioritize software requirements into logical groups with proper identification and priority assignment.",
        "focus_area": "Software requirements organization and prioritization",
        "key_attributes": ["ID", "Object_Type", "ASIL"],
        "llm_focus": "Evaluate organizational structure, logical grouping, prioritization approach, and systematic organization from ID patterns and available data."
    },
    "BP3": {
        "name": "Analyze software requirements",
        "description": "Analyze software requirements for correctness, technical feasibility, and interdependencies to support project management.",
        "focus_area": "Software requirements analysis and feasibility assessment",
        "key_attributes": ["Verification_Strategy", "Verification_Criteria", "Text"],
        "llm_focus": "Assess requirement analysis depth, technical feasibility consideration, interdependency awareness, and verification planning quality."
    },
    "BP4": {
        "name": "Analyze impact on operating environment",
        "description": "Analyze how software requirements impact the operating environment and ensure successful implementation within that environment.",
        "focus_area": "Operating environment impact and implementation feasibility",
        "key_attributes": ["Text", "Verification_Criteria"],
        "llm_focus": "Identify operating environment considerations, implementation constraints, environmental compatibility, and deployment factors from requirement content."
    }
}

# Valid object status values
VALID_OBJECT_STATUS = ["new", "in review", "in progress", "accepted","approved", "rejected"]

STATUS_SCORES = {
    "Compliant": 100.0,
    "Relation-based": 80.0,
    "Partial": 60.0,
    "Non-compliant": 30.0,
    "Unknown": 50.0,
}

def load_analysis_specific_config(base_config_path: str, analysis_type: str = None) -> Dict[str, Any]:
    """Load configuration with analysis-type-specific support"""
    
    config_data = {}
    
    # Try to load base config first
    if os.path.exists(base_config_path):
        try:
            with open(base_config_path, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
                logger.info(f"Loaded base config from {base_config_path}")
        except (json.JSONDecodeError, IOError) as e:
            logger.warning(f"Could not load base config {base_config_path}: {e}")
    
    # If analysis type is specified, try to load type-specific config
    if analysis_type:
        # Determine type-specific config file name
        base_dir = os.path.dirname(base_config_path) or "."
        base_name = os.path.splitext(os.path.basename(base_config_path))[0]
        
        type_config_path = os.path.join(base_dir, f"{base_name}_{analysis_type.lower()}.json")
        
        if os.path.exists(type_config_path):
            try:
                with open(type_config_path, 'r', encoding='utf-8') as f:
                    type_config = json.load(f)
                    # Merge type-specific config with base config (type-specific takes precedence)
                    config_data.update(type_config)
                    logger.info(f"Loaded {analysis_type}-specific config from {type_config_path}")
            except (json.JSONDecodeError, IOError) as e:
                logger.warning(f"Could not load type-specific config {type_config_path}: {e}")
        else:
            logger.info(f"No {analysis_type}-specific config found at {type_config_path}, using base config")
    
    return config_data

def create_example_configs():
    """Create example configuration files for both analysis types"""
    
    # Base configuration
    base_config = {
        "model": "mistral:7b",
        "endpoint": "http://98.86.11.72:11434"
    }
    
    # SYS2-specific configuration (no input file - always from user)
    sys2_config = {
        "model": "mistral:7b",
        "endpoint": "http://98.86.11.72:11434",
        "output": "sys2_analysis_results.xlsx",
        "type": "SYS2",
        "description": "Configuration for SYS.2 System Requirements Analysis - Input file will be requested from user"
    }
    
    # SWE1-specific configuration (no input file - always from user)
    swe1_config = {
        "model": "mistral:7b", 
        "endpoint": "http://98.86.11.72:11434",
        "output": "swe1_analysis_results.xlsx", 
        "type": "SWE1",
        "description": "Configuration for SWE.1 Software Requirements Analysis - Input file will be requested from user"
    }
    
    # Write example config files
    configs_to_create = [
        ("config.json", base_config),
        ("config_sys2.json", sys2_config),
        ("config_swe1.json", swe1_config)
    ]
    
    created_files = []
    for filename, config in configs_to_create:
        if not os.path.exists(filename):
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(config, f, indent=4)
                created_files.append(filename)
            except IOError as e:
                logger.warning(f"Could not create example config {filename}: {e}")
    
    return created_files

def load_csv_with_encoding_detection(file_path: str) -> pd.DataFrame:
    """Load CSV file with automatic encoding detection and error handling"""
    # List of common encodings to try
    encodings_to_try = [
        'utf-8',
        'utf-8-sig',  # UTF-8 with BOM
        'iso-8859-1',  # Latin-1
        'windows-1252',  # Windows encoding
        'cp1252',  # Another Windows encoding
        'latin1',
        'ascii'
    ]
    
    logger.info(f"Attempting to load CSV file: {file_path}")
    
    for encoding in encodings_to_try:
        try:
            logger.info(f"Trying encoding: {encoding}")
            df = pd.read_csv(file_path, encoding=encoding)
            logger.info(f"Successfully loaded CSV with encoding: {encoding}")
            logger.info(f"CSV shape: {df.shape}")
            logger.info(f"Columns: {list(df.columns)}")
            return df
        except UnicodeDecodeError as e:
            logger.warning(f"Failed with encoding {encoding}: {e}")
            continue
        except Exception as e:
            logger.error(f"Error reading CSV with encoding {encoding}: {e}")
            continue
    
    # If all encodings fail, try with error handling
    try:
        logger.info("Trying UTF-8 with error handling (replacing invalid characters)")
        df = pd.read_csv(file_path, encoding='utf-8', encoding_errors='replace')
        logger.warning("Loaded CSV with UTF-8 and replaced invalid characters - some data may be corrupted")
        return df
    except Exception as e:
        raise Exception(f"Could not read CSV file with any encoding. Last error: {e}")

def validate_csv_structure(df: pd.DataFrame) -> bool:
    """Validate that the CSV has the expected structure for requirements analysis"""
    required_columns = ['ID', 'Text']
    missing_columns = [col for col in required_columns if col not in df.columns]
    
    if missing_columns:
        print(f"Warning: Missing required columns: {missing_columns}")
        print(f"Available columns: {list(df.columns)}")
        return False
    
    if df.empty:
        print("Error: CSV file is empty")
        return False
    
    return True

def get_user_input():
    """Interactive menu to get user preferences - now always asks for input file"""
    print("\n" + "="*60)
    print("       UNIFIED ASPICE REQUIREMENTS ANALYZER")
    print("="*60)
    print()
    print("Select Analysis Mode:")
    print("1. SWE.1 - Software Requirements Analysis (BP1-BP4)")
    print("2. SYS.2 - System Requirements Analysis (BP1-BP4)")
    print()
    
    # Get analysis type choice
    while True:
        try:
            choice = input("Enter your choice (1 or 2): ").strip()
            if choice == "1":
                analysis_type = "SWE1"
                print(f"\nSelected: SWE.1 Software Requirements Analysis")
                break
            elif choice == "2":
                analysis_type = "SYS2"
                print(f"\nSelected: SYS.2 System Requirements Analysis")
                break
            else:
                print("Invalid choice. Please enter 1 or 2.")
        except KeyboardInterrupt:
            print("\nOperation cancelled by user.")
            exit(0)
        except EOFError:
            print("\nNo input received.")
            exit(0)
    
    print()
    
    # Get input file path - always ask user
    while True:
        try:
            input_path = input("Enter the path to your requirements CSV file: ").strip()
            
            # Remove quotes if user wrapped path in quotes
            if input_path.startswith('"') and input_path.endswith('"'):
                input_path = input_path[1:-1]
            elif input_path.startswith("'") and input_path.endswith("'"):
                input_path = input_path[1:-1]
            
            if not input_path:
                print("Please enter a valid file path.")
                continue
                
            # Check if file exists
            if not os.path.exists(input_path):
                print(f"File not found: {input_path}")
                print("Please check the path and try again.")
                continue
                
            # Check if it's a CSV file
            if not input_path.lower().endswith('.csv'):
                print("Warning: File doesn't have .csv extension. Continuing anyway...")
            
            print(f"File found: {input_path}")
            break
            
        except KeyboardInterrupt:
            print("\nOperation cancelled by user.")
            exit(0)
        except EOFError:
            print("\nNo input received.")
            exit(0)
    
    return analysis_type, input_path

def get_input_file_from_user():
    """Separate function to get input file path from user"""
    while True:
        try:
            input_path = input("Enter the path to your requirements CSV file: ").strip()
            
            # Remove quotes if user wrapped path in quotes
            if input_path.startswith('"') and input_path.endswith('"'):
                input_path = input_path[1:-1]
            elif input_path.startswith("'") and input_path.endswith("'"):
                input_path = input_path[1:-1]
            
            if not input_path:
                print("Please enter a valid file path.")
                continue
                
            # Check if file exists
            if not os.path.exists(input_path):
                print(f"File not found: {input_path}")
                print("Please check the path and try again.")
                continue
                
            # Check if it's a CSV file
            if not input_path.lower().endswith('.csv'):
                print("Warning: File doesn't have .csv extension. Continuing anyway...")
            
            print(f"File found: {input_path}")
            return input_path
            
        except KeyboardInterrupt:
            print("\nOperation cancelled by user.")
            exit(0)
        except EOFError:
            print("\nNo input received.")
            exit(0)

class UnifiedASPICEChecker:
    """Unified ASPICE compliance checker supporting both SYS.2 and SWE.1"""
    
    def __init__(self, analysis_type: str = "SYS2", model: str = "mistral:7b", endpoint: str = "http://98.86.11.72:11434"):
        self.analysis_type = analysis_type.upper()
        self.model = model
        self.endpoint = endpoint
        self.structural_analysis = None
        
        # Set configuration based on analysis type
        if self.analysis_type == "SYS2":
            self.base_practices = SYS2_BASE_PRACTICES
            self.mandatory_attributes = ["ID", "Text", "Object_Type", "Object_Status", 
                                       "Verification_Criteria", "Verification_Strategy", "System_Feature"]
            self.analysis_title = "SYS.2 System Requirements"
        elif self.analysis_type == "SWE1":
            self.base_practices = SWE1_BASE_PRACTICES
            # Updated SWE1 attributes: removed Priority, added ASIL
            self.mandatory_attributes = ["ID", "Text", "Object_Type", "Object_Status", 
                                       "Verification_Criteria", "Verification_Strategy", "ASIL"]
            self.analysis_title = "SWE.1 Software Requirements"
        else:
            raise ValueError("Analysis type must be 'SYS2' or 'SWE1'")
        
        logger.info(f"Initialized {self.analysis_title} analyzer")
        
    def analyze_file_structure(self, df: pd.DataFrame) -> StructuralAnalysis:
        """Comprehensive structural analysis of requirements file"""
        logger.info(f"Analyzing {self.analysis_title} file structure...")
        
        # Detect groups through multiple methods
        groups_detected = []
        has_groups = False
        
        ##V.2
        # # Method 1: Look for group indicators in ID patterns
        # if 'ID' in df.columns:
        #     id_patterns = df['ID'].str.extract(r'([A-Z]+)[_-]?(\d+)?', expand=True)
        #     if not id_patterns.empty and not id_patterns[0].isna().all():
        #         prefixes = id_patterns[0].value_counts()
        #         if len(prefixes) > 1:
        #             has_groups = True
        #             groups_detected = prefixes.index.tolist()
        
        # Method 2: Look for explicit group types
        # if 'Object_Type' in df.columns:
        #     group_types = df[df['Object_Type'].str.contains('group|Group|GROUP', case=False, na=False)]
        #     if not group_types.empty:
        #         has_groups = True
        #         groups_detected.extend(group_types['ID'].tolist())
        
        # # Method 3: Look for hierarchical patterns in text or structure
        # if 'Text' in df.columns:
        #     if self.analysis_type == "SYS2":
        #         feature_keywords = ['feature', 'module', 'subsystem', 'component']
        #     else:  # SWE1
        #         feature_keywords = ['feature', 'module', 'subsystem', 'component', 'interface', 'function']
            
        #     for text in df['Text'].fillna(''):
        #         if any(keyword in text.lower() for keyword in feature_keywords):
        #             has_groups = True
        #             break
        
        # Analyze functional vs non-functional classification
        if self.analysis_type == "SYS2":
            functional_keywords = ['shall', 'must', 'will', 'function', 'process', 'calculate', 'display', 'control', 'provide', 'enable']
            non_functional_keywords = ['performance', 'reliability', 'usability', 'efficiency', 'maintainability', 'portability', 'security', 'scalability']
        else:  # SWE1
            functional_keywords = ['shall', 'must', 'will', 'function', 'process', 'calculate', 'display', 'control', 'provide', 'enable', 'execute', 'perform']
            non_functional_keywords = ['performance', 'reliability', 'usability', 'efficiency', 'maintainability', 'portability', 'security', 'scalability', 'response time', 'throughput']
        
        functional_count = 0
        non_functional_count = 0
        undefined_count = 0
        atomic_count = 0
        
        for _, row in df.iterrows():
            if pd.isna(row.get('Text', '')):
                undefined_count += 1
                continue
                
            text_lower = str(row['Text']).lower()
            obj_type_lower = str(row.get('Object_Type', '')).lower()
            
            # Skip group containers
            if obj_type_lower == 'group':
                continue
                
            # Check explicit object type classification
            if 'functional' in obj_type_lower and 'non-functional' not in obj_type_lower:
                functional_count += 1
            elif 'non-functional' in obj_type_lower:
                non_functional_count += 1
            # Check text content
            elif any(keyword in text_lower for keyword in functional_keywords):
                if not any(keyword in text_lower for keyword in non_functional_keywords):
                    functional_count += 1
                else:
                    non_functional_count += 1
            elif any(keyword in text_lower for keyword in non_functional_keywords):
                non_functional_count += 1
            else:
                undefined_count += 1
            
            # Check if requirement appears atomic (single responsibility)
            text_sentences = text_lower.split('.')
            if len(text_sentences) <= 2 and text_lower.count(' and ') <= 1 and ' or ' not in text_lower:
                atomic_count += 1
        
        # Handle case where Object_Type column might not exist
        if 'Object_Type' in df.columns:
            total_requirements = len(df) - len(df[df['Object_Type'].astype(str).str.lower() == 'group'])
        else:
            total_requirements = len(df)
        
        functional_coverage = (functional_count + non_functional_count) / total_requirements if total_requirements > 0 else 0
        atomic_requirements_ratio = atomic_count / total_requirements if total_requirements > 0 else 0
        
        # Calculate structural completeness score
        group_score = 1.0 if has_groups else 0.2
        classification_score = functional_coverage
        atomic_score = atomic_requirements_ratio if self.analysis_type == "SWE1" else 1.0
        structural_completeness_score = (group_score + classification_score + atomic_score) / 3
        
        return StructuralAnalysis(
            has_groups=has_groups,
            group_count=len(set(groups_detected)),
            groups_detected=list(set(groups_detected)),
            functional_requirements=functional_count,
            non_functional_requirements=non_functional_count,
            undefined_type_requirements=undefined_count,
            total_requirements=total_requirements,
            functional_coverage=functional_coverage,
            structural_completeness_score=structural_completeness_score,
            atomic_requirements_ratio=atomic_requirements_ratio
        )
    
    def get_missing_attributes(self, row: pd.Series) -> List[str]:
        """Identify missing mandatory attributes for a requirement"""
        missing = []
        for attr in self.mandatory_attributes:
            value = row.get(attr)
            if value is None or (isinstance(value, float) and pd.isna(value)) or str(value).strip().lower() in ['', 'nan', 'tbd', 'todo']:
                missing.append(attr)
        return sorted(set(missing))
    
    def validate_object_status(self, status: str) -> bool:
        """Validate if object status is one of the allowed values"""
        if not status or pd.isna(status):
            return False
        return str(status).lower().strip() in VALID_OBJECT_STATUS
    
    def assess_requirement_quality(self, row: pd.Series) -> Dict[str, bool]:
        """Assess requirement quality against the 5-point checklist"""
        text = str(row.get('Text', '')).lower()
        ver_strategy = str(row.get('Verification_Strategy', ''))
        ver_criteria = str(row.get('Verification_Criteria', ''))
        
        quality_results = {}
        
        # 1. Complete - it fully describes the functionality
        complete_indicators = ['shall', 'must', 'will', 'when', 'where', 'how', 'what']
        has_complete_structure = any(indicator in text for indicator in complete_indicators)
        has_sufficient_detail = len(text.strip()) > 20
        quality_results['Complete'] = has_complete_structure and has_sufficient_detail
        
        # 2. Correct - it accurately describes the functionality
        ambiguous_words = ['appropriate', 'suitable', 'adequate', 'reasonable', 'normal', 'etc', 'and so on']
        contradictory_words = ['but', 'however', 'except when', 'unless']
        is_specific = not any(word in text for word in ambiguous_words)
        not_contradictory = not any(word in text for word in contradictory_words)
        quality_results['Correct'] = is_specific and not_contradictory and len(text.strip()) > 10
        
        # 3. Feasible - the requirement can be implemented
        impossible_words = ['impossible', 'cannot', 'never', 'always 100%', 'instantaneous', 'infinite']
        unrealistic_phrases = ['zero defects', 'perfect', 'never fail', 'always work']
        is_realistic = not any(word in text for word in impossible_words)
        not_unrealistic = not any(phrase in text for phrase in unrealistic_phrases)
        quality_results['Feasible'] = is_realistic and not_unrealistic
        
        # 4. Verifiable - the system can be tested to confirm that it meets the requirement
        has_verification_strategy = ver_strategy.lower() not in ['', 'nan', 'tbd', 'none']
        has_verification_criteria = ver_criteria.lower() not in ['', 'nan', 'tbd', 'none']
        has_measurable_criteria = any(word in text for word in ['measure', 'value', 'time', 'count', 'level', 'rate', 'percentage'])
        quality_results['Verifiable'] = has_verification_strategy or has_verification_criteria or has_measurable_criteria
        
        # 5. Atomic - can't describe more than one function and can't be driven for more than requirement
        multiple_requirements = text.count(' and ') > 2 or text.count(' or ') > 1
        multiple_actions = sum(1 for word in ['shall', 'must', 'will'] if text.count(word) > 1)
        is_single_function = not multiple_requirements and multiple_actions <= 1
        quality_results['Atomic'] = is_single_function and len(text.split('.')) <= 2
        
        return quality_results
    
    def check_bp_compliance(self, bp_id: str, row: pd.Series, structural_analysis: StructuralAnalysis, missing_attrs: List[str]) -> float:
        """Check compliance for specific base practice"""
        if self.analysis_type == "SYS2":
            return self._check_sys2_bp_compliance(bp_id, row, structural_analysis, missing_attrs)
        else:
            return self._check_swe1_bp_compliance(bp_id, row, structural_analysis, missing_attrs)
    
    def _check_sys2_bp_compliance(self, bp_id: str, row: pd.Series, structural_analysis: StructuralAnalysis, missing_attrs: List[str]) -> float:
        """SYS.2 specific compliance checking"""
        if bp_id == "BP1":
            score = 30
            req_id = str(row.get('ID', ''))
            if req_id and req_id.lower() not in ['', 'nan', 'none']:
                score += 20
            text = str(row.get('Text', ''))
            if text and len(text.strip()) > 5:
                score += 30
            obj_type = str(row.get('Object_Type', ''))
            if obj_type and obj_type.lower() not in ['', 'nan', 'none']:
                score += 20
            return max(0, min(100, score))
            
        elif bp_id == "BP2":
            base_score = 40
            if structural_analysis.has_groups:
                base_score += 30
            req_id = str(row.get('ID', ''))
            if re.match(r'^[A-Z]+[_-]?\d+', req_id):
                base_score += 20
            obj_type = str(row.get('Object_Type', ''))
            if obj_type and 'functional' in obj_type.lower():
                base_score += 10
            return max(0, min(100, base_score))
            
        elif bp_id == "BP3":
            score = 0.0
            ver_strategy = str(row.get('Verification_Strategy', ''))
            if ver_strategy and ver_strategy.lower() not in ['', 'nan', 'tbd', 'none']:
                score += 50
            ver_criteria = str(row.get('Verification_Criteria', ''))
            if ver_criteria and ver_criteria.lower() not in ['', 'nan', 'tbd', 'none']:
                score += 50
            return max(0, min(100, score))
            
        elif bp_id == "BP4":
            score = 40
            text = str(row.get('Text', '')).lower()
            context_keywords = ['system', 'interface', 'environment', 'external']
            if any(keyword in text for keyword in context_keywords):
                score += 20
            system_feature = str(row.get('System_Feature', ''))
            if system_feature and system_feature.lower() not in ['', 'nan', 'tbd', 'none']:
                score += 25
            ver_criteria = str(row.get('Verification_Criteria', ''))
            if ver_criteria and len(ver_criteria.strip()) > 10:
                score += 15
            return max(0, min(100, score))
        
        return 50.0  # Default
    
    def _check_swe1_bp_compliance(self, bp_id: str, row: pd.Series, structural_analysis: StructuralAnalysis, missing_attrs: List[str]) -> float:
        """SWE.1 specific compliance checking (BP1-BP4 only)"""
        if bp_id == "BP1":
            score = 60
            req_id = str(row.get('ID', ''))
            if req_id and req_id.lower() not in ['', 'nan', 'none']:
                score += 20
            text = str(row.get('Text', ''))
            if text and len(text.strip()) > 5:
                score += 20
            return max(0, min(100, score))
            
        elif bp_id == "BP2":
            score = 60
            if structural_analysis.has_groups:
                score += 30
            else:
                score += 15
            req_id = str(row.get('ID', ''))
            if req_id and len(req_id.strip()) > 0:
                score += 10
            return max(0, min(100, score))
            
        elif bp_id == "BP3":
            score = 70
            ver_strategy = str(row.get('Verification_Strategy', ''))
            if ver_strategy and ver_strategy.lower() not in ['', 'nan', 'tbd', 'none']:
                score += 15
            ver_criteria = str(row.get('Verification_Criteria', ''))
            if ver_criteria and ver_criteria.lower() not in ['', 'nan', 'tbd', 'none']:
                score += 15
            return max(0, min(100, score))
            
        elif bp_id == "BP4":
            score = 75
            text = str(row.get('Text', '')).lower()
            env_keywords = ['environment', 'operating', 'platform', 'deployment', 'runtime', 'system']
            if any(keyword in text for keyword in env_keywords):
                score += 25
            return max(0, min(100, score))
        
        return 50.0  # Default
    
    def call_llm(self, prompt: str) -> str:
        """Send prompt to LLM and return response"""
        url = f"{self.endpoint}/api/generate"
        payload = {
            "model": self.model,
            "prompt": prompt,
            "stream": False,
        }
        try:
            response = requests.post(url, json=payload, timeout=120)
            response.raise_for_status()
            data = response.json()
            return data.get('response', data.get('content', ''))
        except Exception as exc:
            logger.error(f"LLM call failed: {exc}")
            logger.warning("Falling back to rule-based analysis only")
            return ""
    
    def build_enhanced_prompt(self, row: pd.Series, missing_attrs: List[str]) -> str:
        """Build LLM-focused prompt for comprehensive ASPICE assessment"""
        req_id = row.get('ID', '<unknown>')
        req_text = row.get('Text', '<no text>')
        
        prompt_lines = []
        prompt_lines.append(f"ASPICE {self.analysis_type} Compliance Analysis")
        prompt_lines.append(f"Requirement ID: {req_id}")
        prompt_lines.append(f"Requirement Text: {req_text}")
        
        # Add available attributes for context
        prompt_lines.append(f"\nAvailable {self.analysis_type} Requirement Data:")
        available_attrs = ["Object_Type", "Object_Status", "Verification_Strategy", "Verification_Criteria"]
        if self.analysis_type == "SYS2":
            available_attrs.append("System_Feature")
        else:  # SWE1
            available_attrs.append("ASIL")
        
        for attr in available_attrs:
            value = row.get(attr, 'Not provided')
            if value and str(value).lower() not in ['nan', '']:
                prompt_lines.append(f"- {attr}: {value}")
        
        # ASPICE Base Practice definitions with LLM focus
        prompt_lines.append(f"\nASPICE {self.analysis_type} Base Practices - Focus on LLM Assessment:")
        for bp_id, bp_data in self.base_practices.items():
            prompt_lines.append(f"\n{bp_id}: {bp_data['description']}")
            if 'llm_focus' in bp_data:
                prompt_lines.append(f"LLM Assessment Focus: {bp_data['llm_focus']}")
            else:
                logger.warning(f"Missing llm_focus for {bp_id}. Available keys: {list(bp_data.keys())}")
                prompt_lines.append(f"LLM Assessment Focus: Assess {bp_data.get('focus_area', 'this base practice')} using available data.")
        
        # Structural context if available
        if self.structural_analysis:
            prompt_lines.append(f"\nFile-Level Context:")
            prompt_lines.append(f"- Requirements are {'organized in groups' if self.structural_analysis.has_groups else 'not grouped systematically'}")
            prompt_lines.append(f"- Functional classification coverage: {self.structural_analysis.functional_coverage:.1%}")
            if self.analysis_type == "SWE1":
                prompt_lines.append(f"- Atomic requirements ratio: {self.structural_analysis.atomic_requirements_ratio:.1%}")
        
        # Missing data context
        if missing_attrs:
            prompt_lines.append(f"\nNote: Some expected attributes are missing: {', '.join(missing_attrs)}")
            prompt_lines.append("Assess compliance based on available data and infer where reasonable.")
        
        # Enhanced LLM instructions
        analysis_specific_instructions = ""
        if self.analysis_type == "SYS2":
            analysis_specific_instructions = """
Key Assessment Principles for System Requirements:
- BP1: Focus on requirement quality, clarity, and completeness from the text content
- BP2: Assess organizational structure and systematic approach from ID patterns and classification
- BP3: Evaluate verification planning and analysis quality from available verification data
- BP4: Look for system context, interface, environmental considerations, and system feature mapping

System-Specific Considerations:
- Does the requirement derive from stakeholder needs?
- Is the requirement at the appropriate system level?
- Are interfaces and environmental factors addressed?
- Is system feature mapping defined?"""
        else:  # SWE1
            analysis_specific_instructions = """
Key Assessment Principles for Software Requirements:
- BP1: Focus on software requirement specification quality and alignment with system requirements
- BP2: Assess organizational structure, prioritization, and atomic decomposition of software requirements
- BP3: Evaluate analysis depth, technical feasibility, and interdependency consideration for software
- BP4: Look for operating environment impact, deployment considerations, and implementation feasibility

Software-Specific Considerations:
- Does the requirement properly derive from system requirements?
- Is the requirement implementable in software?
- Are technical constraints and dependencies considered?
- Is the requirement atomic ?"""
        
        prompt_lines.append(f"""
INSTRUCTIONS FOR AI ASSESSMENT:
You are an ASPICE {self.analysis_type} expert evaluating this requirement. Focus on intelligent analysis.

For each Base Practice, provide:
1. STATUS: Compliant/Partial/Non-compliant/Relation-based/Unknown
2. RATIONALE: Your expert reasoning (2-3 sentences)
3. SUGGESTION: Specific improvement recommendation if needed

Additionally, assess the requirement quality checklist:
4. QUALITY_CHECKLIST: For each quality attribute, provide true/false:
   - Complete: Does it fully describe the functionality?
   - Correct: Does it accurately describe the functionality?
   - Feasible: Can the requirement be implemented?
   - Verifiable: Can the system be tested to confirm it meets the requirement?
   - Atomic: Does it describe only one function and cannot be split further?

{analysis_specific_instructions}

Return ONLY valid JSON with all base practices and quality checklist: {{"BP1": {{"status": "...", "rationale": "...", "suggestion": "..."}}, ..., "QUALITY_CHECKLIST": {{"Complete": true/false, "Correct": true/false, "Feasible": true/false, "Verifiable": true/false, "Atomic": true/false}}}}
""")
        
        return '\n'.join(prompt_lines)
    
    def analyze_requirement(self, row: pd.Series, structural_analysis: StructuralAnalysis) -> RequirementAnalysis:
        """Comprehensive requirement analysis combining rule-based and LLM approaches"""
        
        req_id = str(row.get('ID', ''))
        missing_attrs = self.get_missing_attributes(row)
        
        # Skip group containers
        if str(row.get('Object_Type', '')).strip().lower() == 'group':
            return RequirementAnalysis(
                requirement_id=req_id,
                is_compliant=True,
                overall_confidence=100.0,
                bp_scores={},
                violated_practices=[],
                recommendations=[],
                # is_consistent=True,
                # is_testable=True,
                missing_attributes=[],
                structural_impact={"note": "Group container - not evaluated"},
                quality_checklist={"Complete": True, "Correct": True, "Feasible": True, "Verifiable": True, "Atomic": True}
            )
        
        # Rule-based BP scoring
        rule_based_scores = {}
        for bp_id in self.base_practices.keys():
            rule_based_scores[bp_id] = self.check_bp_compliance(bp_id, row, structural_analysis, missing_attrs)
        
        # LLM analysis for validation and additional insights
        prompt = self.build_enhanced_prompt(row, missing_attrs)
        llm_response = self.call_llm(prompt)
        
        # Parse LLM response
        llm_scores = {}
        quality_checklist_llm = {}
        try:
            if llm_response.strip():
                llm_parsed = json.loads(llm_response)
                for bp_id in self.base_practices.keys():
                    bp_data = llm_parsed.get(bp_id, {})
                    status = bp_data.get('status', 'Unknown')
                    llm_scores[bp_id] = STATUS_SCORES.get(status, 50.0)
                
                quality_checklist_llm = llm_parsed.get('QUALITY_CHECKLIST', {})
            else:
                llm_scores = rule_based_scores.copy()
        except json.JSONDecodeError:
            llm_scores = rule_based_scores.copy()
            quality_checklist_llm = {}
        
        # Combine rule-based and LLM scores with LLM preference
        final_scores = {}
        for bp_id in self.base_practices.keys():
            rule_score = rule_based_scores[bp_id]
            llm_score = llm_scores.get(bp_id, rule_score)
            final_scores[bp_id] = (rule_score * 0.3) + (llm_score * 0.7)
        
        # Combine rule-based and LLM quality assessment
        rule_based_quality = self.assess_requirement_quality(row)
        
        final_quality_checklist = {}
        for quality_attr in ['Complete', 'Correct', 'Feasible', 'Verifiable', 'Atomic']:
            llm_result = quality_checklist_llm.get(quality_attr)
            rule_result = rule_based_quality.get(quality_attr, False)
            
            if llm_result is not None:
                final_quality_checklist[quality_attr] = llm_result
            else:
                final_quality_checklist[quality_attr] = rule_result
        
        # Determine violations (threshold: 60% for testing phase)
        violated_practices = [bp_id for bp_id, score in final_scores.items() if score < 60]
        
        # Overall compliance
        overall_confidence = sum(final_scores.values()) / len(final_scores)
        is_compliant = len(violated_practices) == 0 and overall_confidence >= 65
        
        # Consistency and testability checks
        # is_consistent = self._check_consistency(row)
        # is_testable = self._check_testability(row)
        is_atomic = self._check_atomic_requirement(row)
        
        # Generate recommendations
        recommendations = self._generate_recommendations(row, violated_practices, final_scores, missing_attrs)
        
        # Structural impact analysis
        structural_impact = self._analyze_structural_impact(row, structural_analysis)
        
        return RequirementAnalysis(
            requirement_id=req_id,
            is_compliant=is_compliant,
            overall_confidence=overall_confidence,
            bp_scores=final_scores,
            violated_practices=violated_practices,
            recommendations=recommendations,
            # is_consistent=is_consistent,
            # is_testable=is_testable,
            missing_attributes=missing_attrs,
            structural_impact=structural_impact,
            quality_checklist=final_quality_checklist,
            is_atomic=is_atomic
        )
    
    # def _check_consistency(self, row: pd.Series) -> bool:
    #     """Check requirement internal consistency"""
    #     text = str(row.get('Text', '')).lower()
    #     obj_type = str(row.get('Object_Type', '')).lower()
    #     obj_status = str(row.get('Object_Status', ''))
        
    #     # Check type-text alignment
    #     if 'functional' in obj_type and any(word in text for word in ['performance', 'reliability', 'usability']):
    #         return False
        
    #     # Check object status validity
    #     if not self.validate_object_status(obj_status):
    #         return False
        
    #     return True
    
    # def _check_testability(self, row: pd.Series) -> bool:
    #     """Check if requirement is testable"""
    #     ver_strategy = str(row.get('Verification_Strategy', ''))
    #     ver_criteria = str(row.get('Verification_Criteria', ''))
        
    #     return (ver_strategy.lower() not in ['', 'nan', 'tbd', 'none'] and 
    #             ver_criteria.lower() not in ['', 'nan', 'tbd', 'none'])
    
    def _check_atomic_requirement(self, row: pd.Series) -> bool:
        """Check if requirement is atomic (single responsibility)"""
        text = str(row.get('Text', ''))
        
        if not text or len(text.strip()) < 10:
            return False
        
        sentences = text.split('.')
        if len(sentences) > 3:
            return False
        
        conjunctions = [' and ', ' or ', ' but ', ' however ', ' also ']
        conjunction_count = sum(1 for conj in conjunctions if conj in text.lower())
        
        return conjunction_count <= 1
    
    def _generate_recommendations(self, row: pd.Series, violated_practices: List[str], 
                                bp_scores: Dict[str, float], missing_attrs: List[str]) -> List[str]:
        """Generate specific recommendations based on violations"""
        recommendations = []
        
        for bp_id in violated_practices:
            if bp_id == "BP1":
                if not row.get('ID') or str(row.get('ID')).lower() in ['', 'nan']:
                    recommendations.append("Assign unique identifier following project naming convention")
                if missing_attrs:
                    recommendations.append(f"Complete missing attributes: {', '.join(missing_attrs[:3])}")
                text = str(row.get('Text', ''))
                if not text or len(text.strip()) < 15:
                    prefix = "software" if self.analysis_type == "SWE1" else "system"
                    recommendations.append(f"Define clear, detailed {prefix} requirement text with 'shall' statements")
                    
            elif bp_id == "BP2":
                if not self.structural_analysis.has_groups:
                    recommendations.append("Organize requirements into logical groups/features")
                if not self._check_atomic_requirement(row):
                    recommendations.append("Break down complex requirements into atomic, single-responsibility requirements")
                    
            elif bp_id == "BP3":
                if not row.get('Verification_Strategy') or str(row.get('Verification_Strategy')).lower() in ['tbd', 'nan', '']:
                    recommendations.append("Define verification strategy (test/analysis/inspection/demonstration)")
                if not row.get('Verification_Criteria') or str(row.get('Verification_Criteria')).lower() in ['tbd', 'nan', '']:
                    recommendations.append("Specify measurable verification criteria with acceptance criteria")
                    
            elif bp_id == "BP4":
                text = str(row.get('Text', '')).lower()
                if self.analysis_type == "SYS2":
                    if 'interface' not in text and 'environment' not in text:
                        recommendations.append("Consider system interfaces and operating environment impact")
                    system_feature = str(row.get('System_Feature', ''))
                    if not system_feature or system_feature.lower() in ['', 'nan', 'tbd']:
                        recommendations.append("Map requirement to appropriate system feature")
                else:  # SWE1
                    if not any(word in text for word in ['environment', 'platform', 'runtime', 'deployment']):
                        recommendations.append("Consider operating environment constraints and deployment requirements")
        
        # Object status validation
        obj_status = str(row.get('Object_Status', ''))
        if not self.validate_object_status(obj_status):
            recommendations.append(f"Set valid object status: {', '.join(VALID_OBJECT_STATUS)}")
        
        return recommendations[:3]
    
    def _analyze_structural_impact(self, row: pd.Series, structural_analysis: StructuralAnalysis) -> Dict[str, str]:
        """Analyze structural impact on this requirement"""
        impact = {}
        
        if not structural_analysis.has_groups:
            impact["grouping"] = f"No logical grouping detected - impacts {self.analysis_type} BP2 score"
        
        if structural_analysis.functional_coverage < 0.7:
            impact["classification"] = f"Poor functional classification coverage ({structural_analysis.functional_coverage:.1%})"
        
        if self.analysis_type == "SWE1" and structural_analysis.atomic_requirements_ratio < 0.6:
            impact["atomicity"] = f"Low atomic requirements ratio ({structural_analysis.atomic_requirements_ratio:.1%}) - many compound requirements detected"
        
        req_id = str(row.get('ID', ''))
        group_detected = any(group in req_id for group in structural_analysis.groups_detected)
        impact["group_membership"] = f"Group membership: {'Yes' if group_detected else 'No'}"
        
        if self.analysis_type == "SYS2":
            system_feature = str(row.get('System_Feature', ''))
            if not system_feature or system_feature.lower() in ['', 'nan', 'tbd']:
                impact["feature_mapping"] = "No system feature mapping - impacts BP4 score"
        
        return impact
    
    def evaluate_requirements(self, df: pd.DataFrame) -> Tuple[pd.DataFrame, StructuralAnalysis, List[RequirementAnalysis]]:
        """Main evaluation function"""
        logger.info(f"Starting comprehensive {self.analysis_title} evaluation...")
        
        # Perform structural analysis
        self.structural_analysis = self.analyze_file_structure(df)
        
        logger.info(f"Structural Analysis Results:")
        logger.info(f"- Groups detected: {self.structural_analysis.has_groups}")
        logger.info(f"- Functional coverage: {self.structural_analysis.functional_coverage:.1%}")
        if self.analysis_type == "SWE1":
            logger.info(f"- Atomic requirements ratio: {self.structural_analysis.atomic_requirements_ratio:.1%}")
        logger.info(f"- Structural completeness: {self.structural_analysis.structural_completeness_score:.1%}")
        
        # Analyze each requirement
        result_rows = []
        analyses = []
        
        for idx, row in df.iterrows():
            logger.info(f"Analyzing requirement {idx + 1}/{len(df)}: {row.get('ID', 'Unknown')}")
            
            analysis = self.analyze_requirement(row, self.structural_analysis)
            analyses.append(analysis)
            
            # Prepare row for DataFrame
            row_result = row.to_dict()
            
            # Add compliance results
            for bp_id, score in analysis.bp_scores.items():
                row_result[bp_id] = f"{score:.1f}%"
            
            # Add quality and compliance columns in specific order
            # Order: Is_Feasible → Is_Verifiable → Is_Consistent → Is_Testable → Is_Atomic → Is_ASPICE_Compliant → Overall_Confidence
            row_result['Is_Feasible'] = 'YES' if analysis.quality_checklist.get('Feasible', False) else 'NO'
            row_result['Is_Verifiable'] = 'YES' if analysis.quality_checklist.get('Verifiable', False) else 'NO'
            # row_result['Is_Consistent'] = 'YES' if analysis.is_consistent else 'NO'
            # row_result['Is_Testable'] = 'YES' if analysis.is_testable else 'NO'
            row_result['Is_Atomic'] = 'YES' if analysis.is_atomic else 'NO'
            row_result['Is_ASPICE_Compliant'] = 'YES' if analysis.is_compliant else 'NO'
            row_result['Overall_Confidence'] = f"{analysis.overall_confidence:.1f}%"
            row_result['Violated_BPs'] = ', '.join(analysis.violated_practices)
            row_result['Recommendations'] = ' | '.join(analysis.recommendations)
            row_result['Missing_Attributes'] = ', '.join(analysis.missing_attributes)
            
            result_rows.append(row_result)
        
        result_df = pd.DataFrame(result_rows)
        return result_df, self.structural_analysis, analyses
    
    def export_to_excel(self, df: pd.DataFrame, structural_analysis: StructuralAnalysis, 
                       analyses: List[RequirementAnalysis], output_path: str):
        """Export comprehensive results to multi-sheet Excel file"""
        logger.info(f"Exporting {self.analysis_title} results to {output_path}")
        
        wb = Workbook()
        
        # Sheet 1: Main Analysis
        ws_main = wb.active
        ws_main.title = f'{self.analysis_type} Analysis'
        
        # Write main data
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_main.append(r)
        
        # Format main sheet
        self._format_main_sheet(ws_main, df)
        
        # Sheet 2: Summary Statistics
        ws_summary = wb.create_sheet(title='Summary')
        self._create_summary_sheet(ws_summary, analyses, structural_analysis)
        
        # Sheet 3: BP Analysis
        ws_bp = wb.create_sheet(title='BP Analysis')
        self._create_bp_analysis_sheet(ws_bp, analyses)
        
        # Sheet 4: Structural Analysis
        ws_struct = wb.create_sheet(title='Structural Analysis')
        self._create_structural_sheet(ws_struct, structural_analysis)
        
        # Sheet 5: Recommendations
        ws_rec = wb.create_sheet(title='Recommendations')
        self._create_recommendations_sheet(ws_rec, analyses)
        
        # Sheet 6: Quality Checklist
        ws_quality = wb.create_sheet(title='Quality Checklist')
        self._create_quality_checklist_sheet(ws_quality, analyses)
        
        wb.save(output_path)
        logger.info(f"{self.analysis_title} Excel export completed successfully")
    
    def _format_main_sheet(self, ws, df):
        """Apply formatting to main analysis sheet"""
        # Header formatting with different colors for different analysis types
        if self.analysis_type == "SYS2":
            header_fill = PatternFill('solid', fgColor='366092')  # Blue for SYS.2
        else:
            header_fill = PatternFill('solid', fgColor='2E7D32')  # Green for SWE.1
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Conditional formatting for compliance
        green_fill = PatternFill('solid', fgColor='C6EFCE')
        red_fill = PatternFill('solid', fgColor='FFC7CE')
        yellow_fill = PatternFill('solid', fgColor='FFEB9C')
        
        # Find ASPICE compliance column (changed from Is_Compliant)
        compliance_col = None
        for idx, col in enumerate(df.columns):
            if 'Is_ASPICE_Compliant' in col:
                compliance_col = idx + 1
                break
        
        if compliance_col:
            for row in range(2, len(df) + 2):
                cell = ws.cell(row=row, column=compliance_col)
                if cell.value == 'YES':
                    cell.fill = green_fill
                elif cell.value == 'NO':
                    cell.fill = red_fill
        
        # Highlight missing attributes
        missing_col = None
        for idx, col in enumerate(df.columns):
            if 'Missing_Attributes' in col:
                missing_col = idx + 1
                break
        
        if missing_col:
            for row in range(2, len(df) + 2):
                cell = ws.cell(row=row, column=missing_col)
                if cell.value and str(cell.value).strip():
                    cell.fill = yellow_fill
        
        # Highlight quality checklist columns (updated for new column names)
        quality_columns = ['Is_Atomic', 'Is_Feasible', 'Is_Verifiable']
        for quality_col_name in quality_columns:
            quality_col = None
            for idx, col in enumerate(df.columns):
                if quality_col_name in col:
                    quality_col = idx + 1
                    break
            
            if quality_col:
                for row in range(2, len(df) + 2):
                    cell = ws.cell(row=row, column=quality_col)
                    if cell.value == 'YES':
                        cell.fill = green_fill
                    elif cell.value == 'NO':
                        cell.fill = red_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    def _create_summary_sheet(self, ws, analyses, structural_analysis):
        """Create summary statistics sheet"""
        total = len(analyses)
        compliant = sum(1 for a in analyses if a.is_compliant)
        avg_confidence = sum(a.overall_confidence for a in analyses) / total if total > 0 else 0
        atomic_count = sum(1 for a in analyses if a.is_atomic)
        
        data = [
            ['Metric', 'Value', 'Percentage'],
            [f'Total {self.analysis_title} Analyzed', total, '100.0%'],
            ['ASPICE Compliant Requirements', compliant, f'{compliant/total*100:.1f}%' if total > 0 else '0%'],
            ['Average Confidence', f'{avg_confidence:.1f}%', ''],
            ['Atomic Requirements', atomic_count, f'{atomic_count/total*100:.1f}%' if total > 0 else '0%'],
            ['', '', ''],
            ['Structural Metrics', '', ''],
            ['Groups Detected', 'Yes' if structural_analysis.has_groups else 'No', ''],
            ['Functional Coverage', f'{structural_analysis.functional_coverage:.1%}', ''],
            ['Structural Completeness', f'{structural_analysis.structural_completeness_score:.1%}', ''],
        ]
        
        if self.analysis_type == "SWE1":
            data.insert(-3, ['Atomic Requirements Ratio', f'{structural_analysis.atomic_requirements_ratio:.1%}', ''])
        
        for row in data:
            ws.append(row)
        
        # Format header
        for cell in ws[1]:
            cell.font = Font(bold=True)
    
    def _create_bp_analysis_sheet(self, ws, analyses):
        """Create BP-specific analysis sheet"""
        bp_data = [['Base Practice', 'Average Score', 'Violations', 'Focus Area']]
        
        for bp_id, bp_details in self.base_practices.items():
            scores = [a.bp_scores.get(bp_id, 0) for a in analyses if a.bp_scores]
            avg_score = sum(scores) / len(scores) if scores else 0
            violations = sum(1 for a in analyses if bp_id in a.violated_practices)
            
            bp_data.append([
                bp_id,
                f'{avg_score:.1f}%',
                f'{violations}/{len(analyses)}',
                bp_details['focus_area']
            ])
        
        for row in bp_data:
            ws.append(row)
        
        # Format
        for cell in ws[1]:
            cell.font = Font(bold=True)
    
    def _create_structural_sheet(self, ws, structural_analysis):
        """Create structural analysis sheet"""
        data = [
            ['Structural Aspect', 'Result', 'Impact'],
            ['Groups Detected', 'Yes' if structural_analysis.has_groups else 'No', f'Critical for {self.analysis_type} BP2'],
            ['Number of Groups', structural_analysis.group_count, 'Organization quality'],
            ['Functional Requirements', structural_analysis.functional_requirements, 'Classification'],
            ['Non-Functional Requirements', structural_analysis.non_functional_requirements, 'Classification'],
            ['Undefined Classification', structural_analysis.undefined_type_requirements, 'Quality issue'],
            ['Functional Coverage', f'{structural_analysis.functional_coverage:.1%}', f'{self.analysis_type} BP2 impact'],
            ['Structural Completeness', f'{structural_analysis.structural_completeness_score:.1%}', 'Overall quality']
        ]
        
        if self.analysis_type == "SWE1":
            data.insert(-1, ['Atomic Requirements Ratio', f'{structural_analysis.atomic_requirements_ratio:.1%}', 'BP2 decomposition quality'])
        
        for row in data:
            ws.append(row)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
    
    def _create_recommendations_sheet(self, ws, analyses):
        """Create recommendations summary sheet"""
        all_recommendations = []
        for analysis in analyses:
            all_recommendations.extend(analysis.recommendations)
        
        if all_recommendations:
            rec_counts = Counter(all_recommendations)
            data = [['Recommendation', 'Frequency', 'Affected Requirements']]
            
            for rec, count in rec_counts.most_common(10):
                data.append([rec, count, f'{count/len(analyses)*100:.1f}%'])
            
            for row in data:
                ws.append(row)
            
            for cell in ws[1]:
                cell.font = Font(bold=True)
    
    def _create_quality_checklist_sheet(self, ws, analyses):
        """Create quality checklist analysis sheet with updated attributes"""
        # Updated quality attributes (removed Complete and Correct)
        quality_attributes = ['Feasible', 'Verifiable', 'Atomic']
        
        # Calculate statistics for each quality attribute
        data = [['Quality Attribute', 'Pass Count', 'Pass Rate', 'Description']]
        
        quality_descriptions = {
            'Feasible': 'Can be implemented',
            'Verifiable': 'Can be tested to confirm requirement is met',
            'Atomic': 'Describes only one function, cannot be split further'
        }
        
        for attr in quality_attributes:
            pass_count = sum(1 for analysis in analyses if analysis.quality_checklist.get(attr, False))
            pass_rate = pass_count / len(analyses) * 100 if analyses else 0
            description = quality_descriptions.get(attr, '')
            
            data.append([
                attr,
                f'{pass_count}/{len(analyses)}',
                f'{pass_rate:.1f}%',
                description
            ])
        
        # # Add additional attributes from main analysis
        # additional_attrs = ['Consistent', 'Testable']
        # for attr in additional_attrs:
        #     if attr == 'Consistent':
        #         pass_count = sum(1 for analysis in analyses if analysis.is_consistent)
        #         description = 'Requirement is internally consistent'
        #     elif attr == 'Testable':
        #         pass_count = sum(1 for analysis in analyses if analysis.is_testable)
        #         description = 'Has defined verification strategy and criteria'
            
        #     pass_rate = pass_count / len(analyses) * 100 if analyses else 0
            
        #     data.append([
        #         attr,
        #         f'{pass_count}/{len(analyses)}',
        #         f'{pass_rate:.1f}%',
        #         description
        #     ])
        
        # Add overall quality score
        overall_passes = []
        for analysis in analyses:
            individual_score = 0
            individual_score += 1 if analysis.quality_checklist.get('Feasible', False) else 0
            individual_score += 1 if analysis.quality_checklist.get('Verifiable', False) else 0
            individual_score += 1 if analysis.quality_checklist.get('Atomic', False) else 0
            # individual_score += 1 if analysis.is_consistent else 0
            # individual_score += 1 if analysis.is_testable else 0
            overall_passes.append(individual_score)
        
        avg_quality_score = sum(overall_passes) / len(overall_passes) if overall_passes else 0
        max_quality_score = 3  # became 3 attributes total
        
        data.append(['', '', '', ''])
        data.append(['Overall Average', f'{avg_quality_score:.1f}/{max_quality_score}', f'{avg_quality_score/max_quality_score*100:.1f}%', 'Average quality attributes passed per requirement'])
        
        for row in data:
            ws.append(row)
        
        # Format header
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Highlight overall average row + len(additional_attrs)
        if len(data) > len(quality_attributes) + 2:
            for cell in ws[len(data)]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill('solid', fgColor='E6F3FF')


def main():
    """Enhanced main function - always gets input file from user"""
    # Initialize variables to avoid scope issues
    analysis_type = "SYS2"
    input_path = ""
    output_path = ""
    model_name = "mistral:7b"
    endpoint_url = "http://98.86.11.72:11434"
    
    try:
        parser = argparse.ArgumentParser(description="Unified ASPICE Requirements Analyzer (SYS.2 & SWE.1)")
        parser.add_argument('--config', '-c', default='config.json', help='Base configuration file path (type-specific configs will be loaded automatically)')
        parser.add_argument('--output', '-o', help='Output Excel file path')
        parser.add_argument('--model', help='LLM model name (overrides config)')
        parser.add_argument('--endpoint', help='LLM endpoint URL (overrides config)')
        parser.add_argument('--type', '-t', choices=['SYS2', 'SWE1'], 
                           help='Analysis type: SYS2 for System Requirements or SWE1 for Software Requirements')
        parser.add_argument('--interactive', action='store_true', help='Force full interactive mode')
        parser.add_argument('--create-configs', action='store_true', help='Create example configuration files and exit')
        
        args = parser.parse_args()
        
        # Handle config creation request
        if args.create_configs:
            created_files = create_example_configs()
            print("Created example configuration files:")
            for filename in created_files:
                print(f"  - {filename}")
            print("\nEdit these files to customize settings for each analysis type.")
            print("Note: Input files will always be requested from user, not read from config.")
            return
        
        # Determine analysis type
        if args.interactive or not args.type:
            # Full interactive mode - get both analysis type and input file
            analysis_type, input_path = get_user_input()
        else:
            # Partial interactive mode - get analysis type from args, but still ask for input file
            analysis_type = args.type
            print(f"\nAnalysis Type: {analysis_type} ({'Software Requirements' if analysis_type == 'SWE1' else 'System Requirements'})")
            print("Input file will be requested from user input (not config file).")
            input_path = get_input_file_from_user()
        
        # Load analysis-type-specific configuration (excluding input file)
        config_data = load_analysis_specific_config(args.config, analysis_type)
        
        # Remove input from config data if it exists (we always want user input)
        if 'input' in config_data:
            del config_data['input']
            print("Note: Ignoring input file path from config - using user-provided path instead.")
        
        # Log which config files were used
        print(f"\nConfiguration loaded for {analysis_type} analysis")
        if config_data:
            print(f"Settings from config: {', '.join(config_data.keys())}")
        
        # Get other settings from args or config
        model_name = args.model or config_data.get('model', 'mistral:7b')
        endpoint_url = args.endpoint or config_data.get('endpoint', 'http://98.86.11.72:11434')
        output_path = args.output or config_data.get('output')
        
        # Generate output path if not provided
        if not output_path:
            base_name = os.path.splitext(os.path.basename(input_path))[0]
            type_suffix = analysis_type.lower()
            output_path = f"{base_name}_{type_suffix}_analysis.xlsx"
        
        if not output_path.lower().endswith('.xlsx'):
            output_path = f"{output_path}.xlsx"
        
        # Display configuration summary
        print("\n" + "="*60)
        print("CONFIGURATION SUMMARY")
        print("="*60)
        print(f"Analysis Type: {analysis_type} ({'Software Requirements' if analysis_type == 'SWE1' else 'System Requirements'})")
        print(f"Config File: {args.config}")
        if config_data:
            type_config_name = f"{os.path.splitext(args.config)[0]}_{analysis_type.lower()}.json"
            if os.path.exists(type_config_name):
                print(f"Type-Specific Config: {type_config_name}")
        print(f"Input File: {input_path} (user-provided)")
        print(f"Output File: {output_path}")
        print(f"LLM Model: {model_name}")
        print(f"LLM Endpoint: {endpoint_url}")
        if analysis_type == "SWE1":
            print("Note: SWE.1 analysis uses BP1-BP4 only (BP5 and BP6 removed)")
        print("="*60)
        
        # # Confirm before proceeding
        # try:
        #     proceed = input("\nProceed with analysis? (y/n): ").strip().lower()
        #     if proceed not in ['y', 'yes']:
        #         print("Analysis cancelled by user.")
        #         return
        # except (KeyboardInterrupt, EOFError):
        #     print("\nAnalysis cancelled by user.")
        #     return
        
        # Initialize checker and run analysis
        checker = UnifiedASPICEChecker(analysis_type=analysis_type, model=model_name, endpoint=endpoint_url)
        
        logger.info(f"Loading requirements from {input_path}")
        # Use the new encoding-aware CSV loader
        df = load_csv_with_encoding_detection(input_path)
        
        # Validate CSV structure
        if not validate_csv_structure(df):
            raise ValueError("CSV file structure validation failed. Please check the file format and required columns.")
        
        logger.info(f"Starting comprehensive {analysis_type} analysis...")
        result_df, structural_analysis, analyses = checker.evaluate_requirements(df)
        
        logger.info("Exporting results...")
        checker.export_to_excel(result_df, structural_analysis, analyses, output_path)
        
        # Print summary
        total = len(analyses)
        compliant = sum(1 for a in analyses if a.is_compliant)
        avg_confidence = sum(a.overall_confidence for a in analyses) / total if total > 0 else 0
        atomic_count = sum(1 for a in analyses if a.is_atomic)
        
        # Calculate quality checklist statistics (updated for removed attributes)
        quality_stats = {}
        quality_attributes = ['Feasible', 'Verifiable', 'Atomic']
        for attr in quality_attributes:
            pass_count = sum(1 for a in analyses if a.quality_checklist.get(attr, False))
            quality_stats[attr] = (pass_count, pass_count/total*100 if total > 0 else 0)
        
        # Add consistency and testability stats
        # consistent_count = sum(1 for a in analyses if a.is_consistent)
        # testable_count = sum(1 for a in analyses if a.is_testable)
        # quality_stats['Consistent'] = (consistent_count, consistent_count/total*100 if total > 0 else 0)
        # quality_stats['Testable'] = (testable_count, testable_count/total*100 if total > 0 else 0)
        
        print(f"\n{'='*70}")
        print(f"UNIFIED ASPICE {analysis_type} ANALYSIS COMPLETED")
        print(f"{'='*70}")
        print(f"Analysis Type: {checker.analysis_title}")
        if analysis_type == "SWE1":
            print("Note: SWE.1 analysis covers BP1-BP4 (BP5 and BP6 excluded)")
        print(f"Total Requirements Analyzed: {total}")
        print(f"ASPICE Compliant Requirements: {compliant} ({compliant/total*100:.1f}%)" if total > 0 else "No requirements analyzed")
        print(f"Average Confidence Level: {avg_confidence:.1f}%")
        print(f"Atomic Requirements: {atomic_count} ({atomic_count/total*100:.1f}%)" if total > 0 else "")
        print(f"Groups Detected: {'Yes' if structural_analysis.has_groups else 'No'}")
        print(f"Functional Coverage: {structural_analysis.functional_coverage:.1%}")
        print(f"Structural Completeness: {structural_analysis.structural_completeness_score:.1%}")
        if analysis_type == "SWE1":
            print(f"Atomic Requirements Ratio: {structural_analysis.atomic_requirements_ratio:.1%}")
        print(f"\nQuality Assessment Results:")
        for attr, (count, percentage) in quality_stats.items():
            print(f"- {attr}: {count}/{total} ({percentage:.1f}%)")
        print(f"\nResults exported to: {output_path}")
        print("Multi-sheet Excel file includes:")
        print(f"- {analysis_type} Analysis (main results)")
        print("- Summary (statistics)")
        print("- BP Analysis (base practice performance)")
        print("- Structural Analysis (file organization)")
        print("- Recommendations (improvement suggestions)")
        print("- Quality Checklist (requirement quality assessment)")
        print(f"{'='*70}")
        
        input("\nPress Enter to exit...")
            
    except FileNotFoundError:
        print(f"\nError: Could not find the input file: {input_path}")
        print("Please check the file path and try again.")
        input("\nPress Enter to exit...")
    except pd.errors.EmptyDataError:
        print(f"\nError: The input file appears to be empty: {input_path}")
        input("\nPress Enter to exit...")
    except pd.errors.ParserError as e:
        print(f"\nError: Could not parse the CSV file: {input_path}")
        print(f"Parser error: {e}")
        print("Please check that the file is a valid CSV format.")
        input("\nPress Enter to exit...")
    except UnicodeDecodeError as e:
        print(f"\nError: Unicode encoding issue with file: {input_path}")
        print(f"Encoding error: {e}")
        print("This usually happens when the CSV file contains special characters.")
        print("Try saving the file with UTF-8 encoding or use a different CSV file.")
        input("\nPress Enter to exit...")
    except ValueError as e:
        print(f"\nError: Invalid file format or content: {input_path}")
        print(f"Validation error: {e}")
        print("Please ensure the CSV file has the required columns (ID, Text) and proper format.")
        input("\nPress Enter to exit...")
    except Exception as e:
        print(f"\nUnexpected error occurred: {e}")
        logger.error(f"Analysis failed with error: {e}")
        input("\nPress Enter to exit...")
        raise


if __name__ == '__main__':
    main()